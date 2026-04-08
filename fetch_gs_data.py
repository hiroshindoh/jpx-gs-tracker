import json, re, requests
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl

TARGET = "ゴールドマン"
HEADERS = {"User-Agent": "Mozilla/5.0"}

def get(url):
    print("DL: " + url)
    r = requests.get(url, headers=HEADERS, timeout=30)
    if r.status_code == 200:
        return r.content
    print("NG: HTTP " + str(r.status_code))
    return None

def last_friday():
    d = datetime.today()
    days_back = (d.weekday() + 3) % 7 or 7
    return d - timedelta(days=days_back)

def oi_url(dt):
    return ("https://www.jpx.co.jp/automation/markets/derivatives/"
            "open-interest/files/" + dt.strftime("%Y") + "/"
            + dt.strftime("%Y%m%d") + "_nk225op_oi_by_tp.xlsx")

def fut_url(dt):
    return ("https://www.jpx.co.jp/automation/markets/derivatives/"
            "open-interest/files/" + dt.strftime("%Y") + "/"
            + dt.strftime("%Y%m%d") + "_indexfut_oi_by_tp.xlsx")

def vol_urls(dt):
    base = ("https://www.jpx.co.jp/automation/markets/derivatives/"
            "participant-volume/files/daily/" + dt.strftime("%Y%m") + "/"
            + dt.strftime("%Y%m%d"))
    return [
        base + "_volume_by_participant_whole_day.xlsx",
        base + "_volume_by_participant_whole_day_J-NET.xlsx",
        base + "_volume_by_participant_night.xlsx",
        base + "_volume_by_participant_night_J-NET.xlsx",
    ]

def parse_oi(raw, date_str):
    wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    results, strikes = [], set()
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        ps = row[1] if len(row) > 1 else None
        if ps and isinstance(ps, (int, float)):
            strikes.add(int(ps))
            if TARGET in str(row[3] or ""):
                results.append({"date": date_str, "strike": int(ps), "type": "PUT", "side": "Sell", "qty": int(row[4] or 0)})
            if TARGET in str(row[6] or ""):
                results.append({"date": date_str, "strike": int(ps), "type": "PUT", "side": "Buy", "qty": int(row[7] or 0)})
        cs = row[11] if len(row) > 11 else None
        if cs and isinstance(cs, (int, float)):
            strikes.add(int(cs))
            if TARGET in str(row[13] or ""):
                results.append({"date": date_str, "strike": int(cs), "type": "CALL", "side": "Sell", "qty": int(row[14] or 0)})
            if TARGET in str(row[16] or ""):
                results.append({"date": date_str, "strike": int(cs), "type": "CALL", "side": "Buy", "qty": int(row[17] or 0)})
    return results, sorted(strikes)

def parse_fut_oi(raw):
    wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    section = ""
    result = {}
    month_map = {"04月": "2604", "05月": "2605", "06月": "2606", "09月": "2609"}
    section_map = {"日経225先物": "NK225", "日経225mini": "MINI", "TOPIX先物": "TOPIX"}
    for row in rows:
        row = list(row)
        if row[0] and "先物" in str(row[0]):
            for k, v in section_map.items():
                if k in str(row[0]):
                    section = v
                    break
            continue
        if not section or len(row) < 8:
            continue
        month_str = str(row[1] or "")
        month_key = None
        for m, k in month_map.items():
            if m in month_str:
                month_key = k
                break
        if month_key:
            key = section + "_" + month_key
            if TARGET in str(row[3] or ""):
                result[key] = -int(row[4] or 0)
            if TARGET in str(row[6] or ""):
                result[key] = int(row[7] or 0)
        if len(row) > 16:
            month_str2 = str(row[11] or "")
            month_key2 = None
            for m, k in month_map.items():
                if m in month_str2:
                    month_key2 = k
                    break
            if month_key2:
                key2 = section + "_" + month_key2
                if TARGET in str(row[13] or ""):
                    result[key2] = -int(row[14] or 0)
                if TARGET in str(row[16] or ""):
                    result[key2] = int(row[17] or 0)
    return result

def parse_vol(raw, date_str):
    wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    futures_map = {
        "NK225F":  {"2606": "L6", "2609": "L9"},
        "NK225MF": {"2604": "M4", "2605": "M5", "2606": "M6"},
        "TOPIXF":  {"2606": "T6", "2609": "T9"},
    }
    history_day = {}
    opt_list = []
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        if len(row) < 8: continue
        if TARGET not in str(row[5] or ""): continue
        product  = str(row[0] or "")
        contract = str(row[2] or "")
        vol = row[7]
        if isinstance(vol, str) and vol.startswith("="):
            vol = float(vol[1:])
        if not vol: continue
        vol = int(vol)
        if product in futures_map:
            for suffix, key in futures_map[product].items():
                if suffix in contract:
                    history_day[key] = history_day.get(key, 0) + vol
        if product == "NK225E":
            m = re.search(r'([PC])(\d{4})-(\d+)', contract)
            if m:
                opt_type = "CALL" if m.group(1) == "C" else "PUT"
                strike   = int(m.group(3))
                # 同一ストライク・種別があれば加算
                found = False
                for existing in opt_list:
                    if existing["strike"] == strike and existing["type"] == opt_type:
                        existing["volume"] += vol
                        found = True
                        break
                if not found:
                    opt_list.append({
                        "date": date_str, "product": product,
                        "contract": contract, "type": opt_type,
                        "strike": strike, "volume": vol
                    })
    return history_day, opt_list

def main():
    today     = datetime.today()
    today_str = today.strftime("%Y-%m-%d")
    print("今日: " + today_str)

    try:
        with open("gs_data.json", "r", encoding="utf-8") as f:
            existing = json.load(f)
    except Exception:
        existing = {}

    history        = existing.get("history", {})
    oi_history     = existing.get("oi_history", [])
    fut_oi_history = existing.get("fut_oi_history", {})

    # ── オプション建玉残高（週次）──
    friday   = last_friday()
    date_str = friday.strftime("%Y-%m-%d")
    print("建玉残高対象日: " + date_str)
    raw_oi = get(oi_url(friday))
    if raw_oi is None:
        friday2  = friday - timedelta(days=7)
        date_str = friday2.strftime("%Y-%m-%d")
        raw_oi   = get(oi_url(friday2))
    if raw_oi:
        oi_data, all_strikes = parse_oi(raw_oi, date_str)
        for d in oi_data:
            print(str(d["strike"]) + "円 " + d["type"] + " " + d["side"] + ": " + str(d["qty"]) + "枚")
        new_entry = {
            "date": date_str, "gs": oi_data,
            "strikes": {
                "min": min(all_strikes) if all_strikes else 0,
                "max": max(all_strikes) if all_strikes else 0,
                "all": all_strikes
            }
        }
        oi_history = [h for h in oi_history if h["date"] != date_str]
        oi_history.append(new_entry)
        oi_history.sort(key=lambda x: x["date"])
    else:
        oi_data, all_strikes = [], []
        print("オプション建玉残高: 取得失敗")

    # ── 先物建玉残高（週次）──
    raw_fut = get(fut_url(friday))
    if raw_fut is None:
        raw_fut = get(fut_url(friday - timedelta(days=7)))
    if raw_fut:
        fut_data = parse_fut_oi(raw_fut)
        print("先物建玉: " + str(fut_data))
        fut_oi_history[date_str] = fut_data
        keys = sorted(fut_oi_history.keys())
        if len(keys) > 8:
            for old_key in keys[:-8]:
                del fut_oi_history[old_key]
    else:
        print("先物建玉残高: 取得失敗")

    # ── 日次取引量（日中＋J-NET＋ナイト全4ファイルを合算）──
    vol_data, vol_date, history_day = [], "", {}
    for days_back in range(5):
        dt = today - timedelta(days=days_back)
        urls = vol_urls(dt)
        merged_history = {}
        merged_opts    = {}
        any_success    = False
        for url in urls:
            raw_vol = get(url)
            if not raw_vol:
                continue
            any_success = True
            hday, opts = parse_vol(raw_vol, dt.strftime("%Y-%m-%d"))
            # 先物取引量を合算
            for k, v in hday.items():
                merged_history[k] = merged_history.get(k, 0) + v
            # オプション取引量をstrike+typeをキーに合算
            for o in opts:
                key = (o["strike"], o["type"])
                if key in merged_opts:
                    merged_opts[key]["volume"] += o["volume"]
                else:
                    merged_opts[key] = dict(o)
        if any_success:
            vol_date    = dt.strftime("%Y-%m-%d")
            history_day = merged_history
            vol_data    = list(merged_opts.values())
            print("取引量取得（4セッション合算）: " + vol_date)
            for k, v in history_day.items():
                print("[先物] " + k + ": " + str(v) + "枚")
            for o in sorted(vol_data, key=lambda x: x["strike"]):
                print("[OPT] " + o["type"] + " " + str(o["strike"]) + "円: " + str(o["volume"]) + "枚")
            break

    if history_day and vol_date:
        history[vol_date] = history_day

    output = {
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "oi": {
            "date": date_str, "gs": oi_data,
            "strikes": {
                "min": min(all_strikes) if all_strikes else 0,
                "max": max(all_strikes) if all_strikes else 0,
                "all": all_strikes
            }
        },
        "vol":             {"date": vol_date, "gs": vol_data},
        "history":         history,
        "oi_history":      oi_history,
        "fut_oi_history":  fut_oi_history
    }

    with open("gs_data.json", "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print("完了 ✅")

if __name__ == "__main__":
    main()
